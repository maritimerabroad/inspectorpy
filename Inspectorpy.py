import os
import wx
import openpyxl
from openpyxl.cell import get_column_letter
wildcard = "Excel File (*.xlsx)|*.xlsx|" \
            "All files (*.*)|*.*"
wildcard2 = "Text File (*.txt)|*.txt|" \
            "All files (*.*)|*.*"
########################################################################
class MyForm(wx.Frame):

    #----------------------------------------------------------------------
    def __init__(self):
        wx.Frame.__init__(self, None, wx.ID_ANY,
                          "Inspection Converter", size=(350, 175))
        panel = wx.Panel(self, wx.ID_ANY)
        self.currentDirectory = os.getcwd()

        loadFileDlgBtn = wx.Button(panel, label="Load your Inspection Report excel file")
        loadFileDlgBtn.Bind(wx.EVT_BUTTON, self.loadFile)		
		
        saveFileDlgBtn = wx.Button(panel, label="Save the converted .txt file")
        saveFileDlgBtn.Bind(wx.EVT_BUTTON, self.saveFile)
        

        convertAndExit = wx.Button(panel, label="Convert and Quit")
        convertAndExit.Bind(wx.EVT_BUTTON, self.convertAndExit)


        # put the buttons in a sizer
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(loadFileDlgBtn, 0, wx.ALL|wx.CENTER, 5)
        sizer.Add(saveFileDlgBtn, 0, wx.ALL|wx.CENTER, 10)
        sizer.Add(convertAndExit, 0, wx.CENTER|wx.CENTER, 15)
        panel.SetSizer(sizer)

#--------------------------------------------------------------------------------		
		
    def convertAndExit(self, event):
        convert(input_stream)
        #exit()	    
    #----------------------------------------------------------------------
    def loadFile(self, event):

        dlg = wx.FileDialog(
            self, message="Load File....", 
            defaultDir=self.currentDirectory, 
            defaultFile="", wildcard=wildcard,
            )
        if dlg.ShowModal() == wx.ID_OK:
            global input_stream
            input_stream = dlg.GetPath()
            print "You chose the following filename: %s" % input_stream
        dlg.Destroy()
	

    #----------------------------------------------------------------------
    def saveFile(self, event):
        """
        Create and show the Save FileDialog
        """
        dlg = wx.FileDialog(
            self, message="Save file as ...", 
            defaultDir=self.currentDirectory, 
            defaultFile="", wildcard=wildcard2, style=wx.SAVE
            )
        if dlg.ShowModal() == wx.ID_OK:
            global output_stream
            output_stream = dlg.GetPath()
            print "You chose the following filename: %s" % output_stream
        dlg.Destroy()
		


def extract_areas(ws):
    all_rows = ws.iter_rows("a2:a100")
    areas = [row[0].value for row in all_rows if row[0].value]
    return areas

def extract_issues(ws):
    issues = [cell.value for cell in next(ws.rows)[1:] if cell.value]
    return issues

def grab_column(ws, column_number, areas):
    letter = get_column_letter(column_number)
    cell_range = '{0}2:{0}{1}'.format(letter, len(areas)) # generate B2:B33 or so
    notes = [row[0].value for row in ws.iter_rows(cell_range)]
    return notes

def convert(input_stream):
    wb = openpyxl.load_workbook(filename=input_stream, read_only=True)
    sheet = wb['Sheet1']
    ws = wb.active
	
    areas = extract_areas(ws)
    issues = extract_issues(ws)
    #print areas
    #print issues

    for idx, issue in enumerate(issues, 2):
        area_notes = grab_column(ws, idx, areas)
        for area, note in zip(areas, area_notes):
            if note:
                print("- {} {} {}".format(issue, area, note))
                with open(output_stream, 'a') as f:
                    f.write("- {} {} {} \n".format(issue, area, note))
# Run the program
if __name__ == "__main__":
    app = wx.App(False)
    frame = MyForm()
    frame.Show()
    app.MainLoop()
